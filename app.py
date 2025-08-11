# --- app.py (PostgreSQL対応・最終完全版) ---
import psycopg2
from psycopg2.extras import RealDictCursor
import secrets
import os
import re
from flask import Flask, jsonify, request, g, send_file, url_for, send_from_directory
from flask_cors import CORS
from datetime import datetime, timezone, timedelta
import openpyxl
from functools import wraps
import io
from PIL import Image
from werkzeug.utils import secure_filename
from flask_bcrypt import Bcrypt
import jwt 

# --- 定数 ---
UPLOAD_FOLDER = 'uploads'
IMAGES_FOLDER = 'images' 

# --- 環境変数 ---
JWT_SECRET_KEY = os.environ.get("JWT_SECRET_KEY")
if not JWT_SECRET_KEY:
    raise ValueError("Error: JWT_SECRET_KEY must be set as an environment variable.")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(IMAGES_FOLDER, exist_ok=True)

# --- Flaskアプリケーション設定 ---
app = Flask(__name__, static_folder='images', static_url_path='/images')

ALLOWED_ORIGINS = os.environ.get("FRONTEND_URL", "http://127.0.0.1:5000")
CORS(app, resources={r"/api/*": {"origins": ALLOWED_ORIGINS}}) 

app.config['JSON_AS_ASCII'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['IMAGES_FOLDER'] = IMAGES_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
bcrypt = Bcrypt(app)

# --- データベースのマイグレーション ---
def migrate_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor(cursor_factory=RealDictCursor)
        
        migrations = [
            "ALTER TABLE calls ADD COLUMN call_type TEXT NOT NULL DEFAULT 'normal'",
            "ALTER TABLE calls ADD COLUMN status TEXT NOT NULL DEFAULT 'new'",
            "ALTER TABLE categories ADD COLUMN display_order INTEGER NOT NULL DEFAULT 99",
            "ALTER TABLE orders ADD COLUMN paid_at REAL",
            "ALTER TABLE orders ADD COLUMN printed_at REAL",
            "ALTER TABLE order_items ADD COLUMN ready_at REAL",
            "ALTER TABLE products ADD COLUMN name_en TEXT",
            "ALTER TABLE products ADD COLUMN description_en TEXT",
        ]
        
        for migration in migrations:
            try:
                cursor.execute(migration)
                db.commit()
                print(f"Migration successful: {migration}")
            except psycopg2.Error as e:
                db.rollback()
                print(f"Skipping migration (already applied or error): {e}")

# --- データベース初期化 ---
def init_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute('CREATE TABLE IF NOT EXISTS orders (id SERIAL PRIMARY KEY, table_id INTEGER, total_price REAL, status TEXT, created_at REAL NOT NULL, paid_at REAL, printed_at REAL)')
        cursor.execute('CREATE TABLE IF NOT EXISTS order_items (id SERIAL PRIMARY KEY, order_id INTEGER, item_name TEXT, quantity INTEGER, price REAL, item_status TEXT, ready_at REAL, FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE)')
        cursor.execute('CREATE TABLE IF NOT EXISTS table_sessions (id SERIAL PRIMARY KEY, table_id INTEGER NOT NULL, access_token TEXT NOT NULL, status TEXT NOT NULL, created_at REAL NOT NULL, UNIQUE(access_token))')
        cursor.execute('CREATE TABLE IF NOT EXISTS categories (id SERIAL PRIMARY KEY, name_jp TEXT NOT NULL, name_en TEXT, display_order INTEGER NOT NULL DEFAULT 99, UNIQUE(name_jp))')
        cursor.execute('CREATE TABLE IF NOT EXISTS products (id SERIAL PRIMARY KEY, name TEXT NOT NULL, price INTEGER NOT NULL, description TEXT, image_path TEXT, is_sold_out INTEGER NOT NULL DEFAULT 0, name_en TEXT, description_en TEXT, UNIQUE(name))')
        cursor.execute('CREATE TABLE IF NOT EXISTS product_categories (product_id INTEGER NOT NULL, category_id INTEGER NOT NULL, PRIMARY KEY (product_id, category_id), FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE, FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE)')
        cursor.execute('CREATE TABLE IF NOT EXISTS calls (id SERIAL PRIMARY KEY, table_id INTEGER NOT NULL, call_time REAL NOT NULL, call_type TEXT NOT NULL DEFAULT \'normal\', status TEXT NOT NULL DEFAULT \'new\', UNIQUE(table_id))')
        cursor.execute('CREATE TABLE IF NOT EXISTS users (id SERIAL PRIMARY KEY, username TEXT NOT NULL, password_hash TEXT NOT NULL, role TEXT NOT NULL, UNIQUE(username))')
        cursor.execute('CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)')

        default_settings = {
            'opening_message': 'ご来店ありがとうございます！', 'opening_image_path': '', 'opening_image_path_2': '',
            'opening_writing_mode': 'horizontal-tb', 'opening_effect': 'fade', 'opening_duration': '5',
            'store_name': '', 'store_address': '', 'store_tel': '', 'store_receipt_note': '',
            'store_qr_code_path': ''
        }
        for key, value in default_settings.items():
            cursor.execute("INSERT INTO settings (key, value) VALUES (%s, %s) ON CONFLICT (key) DO NOTHING", (key, value))

        cursor.execute("SELECT * FROM users WHERE username = %s", ('admin',))
        if not cursor.fetchone():
            password_hash = bcrypt.generate_password_hash('admin_password').decode('utf-8')
            cursor.execute("INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s) ON CONFLICT (username) DO NOTHING",('admin', password_hash, 'admin'))
        
        cursor.execute("SELECT * FROM users WHERE username = %s", ('staff',))
        if not cursor.fetchone():
            password_hash = bcrypt.generate_password_hash('your_common_password').decode('utf-8')
            cursor.execute("INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s) ON CONFLICT (username) DO NOTHING",('staff', password_hash, 'staff'))
        
        db.commit()

# --- データベース接続ヘルパー ---
def get_db():
    if 'db' not in g:
        db_url = os.environ.get("DATABASE_URL")
        if not db_url:
            raise ValueError("DATABASE_URL environment variable is not set")
        g.db = psycopg2.connect(db_url)
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()

# --- 認証デコレータ ---
def staff_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({"status": "error", "message": "Authorization header missing or invalid"}), 401
        token = auth_header.split(' ')[1]
        try:
            payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=["HS256"])
            g.user_username = payload.get('username')
            g.user_role = payload.get('role')
        except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
            return jsonify({"status": "error", "message": "Invalid or expired token"}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_role_required(f):
    @wraps(f)
    @staff_required
    def decorated_function(*args, **kwargs):
        if g.user_role != 'admin':
            return jsonify({"status": "error", "message": "Admin privileges required"}), 403
        return f(*args, **kwargs)
    return decorated_function

# --- ログインAPI ---
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    if not username or not password: return jsonify({"status": "error", "message": "Username and password required"}), 400
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
    user = cursor.fetchone()
    if user and bcrypt.check_password_hash(user['password_hash'], password):
        payload = {'username': user['username'], 'role': user['role'], 'exp': datetime.now(timezone.utc) + timedelta(hours=8)}
        token = jwt.encode(payload, JWT_SECRET_KEY, algorithm="HS256")
        return jsonify({"status": "success", "token": token, "role": user['role']})
    else:
        return jsonify({"status": "error", "message": "Invalid username or password"}), 401

# --- API: 顧客向け ---

@app.route('/api/get_opening_settings')
def get_opening_settings():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT key, value FROM settings WHERE key LIKE 'opening_%'")
    settings = {row['key']: row['value'] for row in cursor.fetchall()}
    if settings.get('opening_image_path'):
        settings['opening_image_url'] = url_for('static', filename=settings['opening_image_path'], _external=True)
    if settings.get('opening_image_path_2'):
        settings['opening_image_url_2'] = url_for('static', filename=settings['opening_image_path_2'], _external=True)
    
    # ★★★ クレジット情報に加えて、メッセージも確実に渡すように修正 ★★★
    cursor.execute("SELECT value FROM settings WHERE key = 'opening_message'")
    msg_row = cursor.fetchone()
    settings['opening_message'] = msg_row['value'] if msg_row else ''
    settings['credit_text'] = "powered by RISE with Google AI Studio"

    return jsonify(settings)


@app.route('/api/get_public_store_info')
def get_public_store_info():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT value FROM settings WHERE key = 'store_name'")
    store_name_row = cursor.fetchone()
    store_name = store_name_row['value'] if store_name_row and store_name_row['value'] else 'レストラン「My Order LINK」'
    return jsonify({"store_name": store_name})

@app.route('/api/get_products')
def get_products():
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT id, name_jp, name_en FROM categories")
    categories_map = {c['id']: c for c in cursor.fetchall()}
    cursor.execute("SELECT * FROM product_categories")
    product_to_cats = {}
    for row in cursor.fetchall():
        pid, cid = row['product_id'], row['category_id']
        if pid not in product_to_cats:
            product_to_cats[pid] = []
        product_to_cats[pid].append(cid)
    cursor.execute("SELECT * FROM products ORDER BY id")
    products = []
    for p_row in cursor.fetchall():
        cat_ids = product_to_cats.get(p_row['id'], [])
        p_row['categories'] = [categories_map[cid] for cid in cat_ids if cid in categories_map]
        p_row['category'] = " ".join([cat['name_jp'] for cat in p_row['categories']])
        products.append(p_row)
    return jsonify(products)

@app.route('/api/get_order_history/<int:table_id>', methods=['GET'])
def get_order_history(table_id):
    access_token = request.args.get('token')
    if not access_token: return jsonify({"status": "error", "message": "Access token is missing."}), 403
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT id FROM table_sessions WHERE table_id = %s AND access_token = %s AND status = 'active'", (table_id, access_token))
    if not cursor.fetchone(): return jsonify({"status": "error", "message": "Invalid or expired access token."}), 403
    cursor.execute("SELECT o.id as order_id, o.created_at, oi.item_name, oi.quantity, oi.price, oi.item_status FROM orders o JOIN order_items oi ON o.id = oi.order_id WHERE o.table_id = %s AND o.status = 'active' ORDER BY o.created_at ASC, oi.id ASC", (table_id,))
    history_items = cursor.fetchall()
    history = {"items": history_items}
    history["total_price"] = sum(item['price'] * item['quantity'] for item in history['items'])
    return jsonify(history)

@app.route('/api/order', methods=['POST'])
def receive_order():
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    order_data = request.get_json()
    table_id, access_token = order_data.get('tableId'), order_data.get('accessToken')
    if not access_token: return jsonify({"status": "error", "message": "Access token is missing."}), 403
    cursor.execute("SELECT id FROM table_sessions WHERE table_id = %s AND access_token = %s AND status = 'active'", (table_id, access_token))
    if not cursor.fetchone(): return jsonify({"status": "error", "message": "Invalid or expired access token."}), 403
    try:
        cursor.execute("SELECT id FROM orders WHERE table_id = %s AND status = 'active'", (table_id,))
        active_order = cursor.fetchone()
        order_id = active_order['id'] if active_order else None
        if not order_id:
            cursor.execute("INSERT INTO orders (table_id, total_price, status, created_at) VALUES (%s, %s, %s, %s) RETURNING id", (table_id, 0, 'active', datetime.now(timezone.utc).timestamp()))
            order_id = cursor.fetchone()['id']
        for item in order_data.get('items', []):
            cursor.execute("SELECT price FROM products WHERE name = %s AND is_sold_out = 0", (item['name'],))
            db_item = cursor.fetchone()
            if not db_item:
                db.rollback()
                return jsonify({"status": "error", "message": f"「{item['name']}」は現在注文できません。"}), 400
            cursor.execute("SELECT id, quantity FROM order_items WHERE order_id = %s AND item_name = %s AND item_status = 'cooking'", (order_id, item['name']))
            existing_item = cursor.fetchone()
            if existing_item:
                new_quantity = existing_item['quantity'] + item['quantity']
                cursor.execute("UPDATE order_items SET quantity = %s WHERE id = %s", (new_quantity, existing_item['id']))
            else:
                cursor.execute("INSERT INTO order_items (order_id, item_name, quantity, price, item_status) VALUES (%s, %s, %s, %s, %s)", (order_id, item['name'], item['quantity'], db_item['price'], 'cooking'))
        cursor.execute("SELECT SUM(price * quantity) FROM order_items WHERE order_id = %s", (order_id,))
        new_total = cursor.fetchone()['sum'] or 0
        cursor.execute("UPDATE orders SET total_price = %s WHERE id = %s", (new_total, order_id))
        db.commit()
        return jsonify({"status": "success", "orderId": order_id})
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/call', methods=['POST'])
def handle_call():
    data = request.get_json()
    table_id, access_token, call_type = data.get('tableId'), data.get('token'), data.get('call_type', 'normal')
    if call_type not in ['normal', 'checkout']: call_type = 'normal'
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT id FROM table_sessions WHERE table_id = %s AND access_token = %s AND status = 'active'", (table_id, access_token))
    if not cursor.fetchone(): return jsonify({"status": "error", "message": "無効なアクセスです"}), 403
    cursor.execute("""
        INSERT INTO calls (table_id, call_time, call_type, status) 
        VALUES (%s, %s, %s, 'new') 
        ON CONFLICT (table_id) 
        DO UPDATE SET call_time = EXCLUDED.call_time, call_type = EXCLUDED.call_type, status = 'new'
    """, (table_id, datetime.now(timezone.utc).timestamp(), call_type))
    db.commit()
    return jsonify({"status": "success"})

    # --- API: ホール・厨房・レジ向け ---
@app.route('/api/generate_table_token/<int:table_id>', methods=['POST'])
@staff_required
def generate_table_token(table_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("UPDATE table_sessions SET status = 'expired' WHERE table_id = %s AND status = 'active'", (table_id,))
    token = secrets.token_urlsafe(16)
    cursor.execute("INSERT INTO table_sessions (table_id, access_token, status, created_at) VALUES (%s, %s, %s, %s)", (table_id, token, 'active', datetime.now(timezone.utc).timestamp()))
    db.commit()
    return jsonify({"status": "success", "accessToken": token, "tableId": table_id})

@app.route('/api/get_calls')
@staff_required
def get_calls():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT table_id, call_time, call_type FROM calls WHERE status = 'new' ORDER BY call_time ASC")
    return jsonify(cursor.fetchall())

@app.route('/api/resolve_call/<int:table_id>', methods=['POST'])
@staff_required
def resolve_call(table_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT call_type FROM calls WHERE table_id = %s AND status = 'new'", (table_id,))
    call = cursor.fetchone()
    if not call:
        return jsonify({"status": "success", "message": "Call not found or already resolved."})
    if call['call_type'] == 'checkout':
        cursor.execute("UPDATE calls SET status = 'acknowledged' WHERE table_id = %s", (table_id,))
    else:
        cursor.execute("DELETE FROM calls WHERE table_id = %s", (table_id,))
    db.commit()
    return jsonify({"status": "success"})

@app.route('/api/get_all_active_orders')
@staff_required
def get_all_active_orders():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT * FROM orders WHERE status != 'paid' ORDER BY created_at ASC")
    orders = cursor.fetchall()
    for order in orders:
        cursor.execute("SELECT * FROM order_items WHERE order_id = %s", (order['id'],))
        order['items'] = cursor.fetchall()
    return jsonify(orders)

@app.route('/api/update_item_status/<int:item_id>', methods=['POST'])
@staff_required
def update_item_status(item_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    new_status = request.get_json().get('status')
    if new_status == 'ready':
        cursor.execute("UPDATE order_items SET item_status = %s, ready_at = %s WHERE id = %s", (new_status, datetime.now(timezone.utc).timestamp(), item_id))
    else:
        cursor.execute("UPDATE order_items SET item_status = %s WHERE id = %s", (new_status, item_id))
    db.commit()
    return jsonify({"status": "success"})

@app.route('/api/update_item_quantity/<int:item_id>', methods=['POST'])
@staff_required
def update_item_quantity(item_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    new_quantity = request.get_json().get('quantity')
    if not isinstance(new_quantity, int) or new_quantity <= 0: return jsonify({"status": "error", "message": "Invalid quantity"}), 400
    cursor.execute("SELECT order_id FROM order_items WHERE id = %s", (item_id,))
    res = cursor.fetchone()
    if not res: return jsonify({"status": "error", "message": "Item not found"}), 404
    order_id = res['order_id']
    cursor.execute("UPDATE order_items SET quantity = %s WHERE id = %s", (new_quantity, item_id))
    cursor.execute("SELECT SUM(price * quantity) FROM order_items WHERE order_id = %s", (order_id,))
    new_total = cursor.fetchone()['sum'] or 0
    cursor.execute("UPDATE orders SET total_price = %s WHERE id = %s", (new_total, order_id))
    db.commit()
    return jsonify({"status": "success"})

@app.route('/api/cancel_item/<int:item_id>', methods=['POST'])
@staff_required
def cancel_item(item_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT order_id FROM order_items WHERE id = %s", (item_id,))
    result = cursor.fetchone()
    if not result: return jsonify({"status": "error", "message": "Item not found"}), 404
    order_id = result['order_id']
    cursor.execute("DELETE FROM order_items WHERE id = %s", (item_id,))
    cursor.execute("SELECT SUM(price * quantity) FROM order_items WHERE order_id = %s", (order_id,))
    new_total = cursor.fetchone()['sum'] or 0
    cursor.execute("UPDATE orders SET total_price = %s WHERE id = %s", (new_total, order_id))
    cursor.execute("SELECT COUNT(*) FROM order_items WHERE order_id = %s", (order_id,))
    if cursor.fetchone()['count'] == 0: cursor.execute("DELETE FROM orders WHERE id = %s", (order_id,))
    db.commit()
    return jsonify({"status": "success"})

@app.route('/api/get_table_summary')
@staff_required
def get_table_summary():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT table_id, call_type FROM calls")
    calls_map = {row['table_id']: row['call_type'] for row in cursor.fetchall()}
    cursor.execute("SELECT * FROM orders WHERE status != 'paid' ORDER BY table_id, created_at ASC")
    table_summary = {}
    for order_row in cursor.fetchall():
        table_id = order_row['table_id']
        if table_id not in table_summary: 
            table_summary[table_id] = {"table_id": table_id, "orders": [], "grand_total": 0, "call_type": calls_map.get(table_id, None)}
        
        cursor.execute("SELECT * FROM order_items WHERE order_id = %s", (order_row['id'],))
        order_row['items'] = cursor.fetchall()
        table_summary[table_id]['orders'].append(order_row)
        table_summary[table_id]['grand_total'] += order_row.get('total_price', 0)
    return jsonify(list(table_summary.values()))

@app.route('/api/checkout_table/<int:table_id>', methods=['POST'])
@staff_required
def checkout_table(table_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("UPDATE orders SET status = 'paid', paid_at = %s WHERE table_id = %s AND status != 'paid'", (datetime.now(timezone.utc).timestamp(), table_id))
    cursor.execute("UPDATE table_sessions SET status = 'expired' WHERE table_id = %s AND status = 'active'", (table_id,))
    cursor.execute("DELETE FROM calls WHERE table_id = %s", (table_id,))
    db.commit()
    return jsonify({"status": "success"})

@app.route('/api/get_paid_orders')
@staff_required
def get_paid_orders():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    start_of_today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).timestamp()
    cursor.execute("SELECT * FROM orders WHERE status = 'paid' AND created_at >= %s ORDER BY paid_at DESC LIMIT 200", (start_of_today,))
    return jsonify(cursor.fetchall())

@app.route('/api/get_order_for_print/<int:order_id>')
@staff_required
def get_order_for_print(order_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT key, value FROM settings WHERE key LIKE 'store_%'")
    store_info = {row['key']: row['value'] for row in cursor.fetchall()}
    if store_info.get('store_qr_code_path'): store_info['store_qr_code_url'] = url_for('static', filename=store_info['store_qr_code_path'])
    cursor.execute("SELECT * FROM orders WHERE id = %s", (order_id,))
    order = cursor.fetchone()
    if not order: return jsonify({"status": "error", "message": "Order not found"}), 404
    cursor.execute("SELECT * FROM order_items WHERE order_id = %s", (order_id,))
    items = cursor.fetchall()
    cursor.execute("UPDATE orders SET printed_at = %s WHERE id = %s", (datetime.now(timezone.utc).timestamp(), order_id))
    db.commit()
    return jsonify({"order": order, "items": items, "store_info": store_info})

# --- API: 顧客・管理者向け ---
@app.route('/api/get_categories')
def get_categories():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT id, name_jp, name_en, display_order FROM categories ORDER BY display_order, id")
    return jsonify(cursor.fetchall())

    # --- 管理者向けAPI ---
@app.route('/api/admin/get_categories')
@admin_role_required
def admin_get_categories():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT * FROM categories ORDER BY display_order, id")
    return jsonify(cursor.fetchall())

@app.route('/api/admin/add_category', methods=['POST'])
@admin_role_required
def add_category():
    data = request.get_json()
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    try:
        cursor.execute("INSERT INTO categories (name_jp, name_en) VALUES (%s, %s) RETURNING id", (data['name_jp'], data['name_en']))
        new_id = cursor.fetchone()['id']
        db.commit()
        return jsonify({"status": "success", "id": new_id})
    except psycopg2.errors.UniqueViolation:
        db.rollback()
        return jsonify({"status": "error", "message": "そのカテゴリー名は既に使用されています。"}), 409
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/update_category/<int:category_id>', methods=['POST'])
@admin_role_required
def update_category(category_id):
    data = request.get_json()
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    try:
        cursor.execute("UPDATE categories SET name_jp = %s, name_en = %s WHERE id = %s", (data['name_jp'], data['name_en'], category_id))
        db.commit()
        return jsonify({"status": "success"})
    except psycopg2.errors.UniqueViolation:
        db.rollback()
        return jsonify({"status": "error", "message": "そのカテゴリー名は既に使用されています。"}), 409
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/delete_category/<int:category_id>', methods=['POST'])
@admin_role_required
def delete_category(category_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    try:
        cursor.execute("SELECT COUNT(*) FROM product_categories WHERE category_id = %s", (category_id,))
        if cursor.fetchone()['count'] > 0: return jsonify({"status": "error", "message": "このカテゴリーを使用しているメニューが存在するため、削除できません。"}), 400
        cursor.execute("DELETE FROM categories WHERE id = %s", (category_id,))
        db.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/get_store_info')
@admin_role_required
def get_store_info():
    cursor = get_db().cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT key, value FROM settings WHERE key LIKE 'store_%'")
    store_info = {row['key']: row['value'] for row in cursor.fetchall()}
    if store_info.get('store_qr_code_path'): store_info['store_qr_code_url'] = url_for('static', filename=store_info['store_qr_code_path'])
    return jsonify(store_info)

@app.route('/api/admin/update_store_info', methods=['POST'])
@admin_role_required
def update_store_info():
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    try:
        for key in ['store_name', 'store_address', 'store_tel', 'store_receipt_note']:
            if key in request.form:
                cursor.execute("INSERT INTO settings (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", (key, request.form[key]))
        if 'store_qr_code' in request.files and request.files['store_qr_code'].filename != '':
            qr_file = request.files['store_qr_code']
            img = Image.open(qr_file.stream)
            ext = os.path.splitext(secure_filename(qr_file.filename))[1].lower()
            if ext not in ['.png', '.jpg', '.jpeg', '.gif']: raise IOError("Invalid QR code image format")
            filename = f"qr_code{ext}"
            img.save(os.path.join(app.config['IMAGES_FOLDER'], filename))
            cursor.execute("UPDATE settings SET value = %s WHERE key = 'store_qr_code_path'", (filename,))
        db.commit()
        return jsonify({"status": "success", "message": "店舗情報を更新しました。"})
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/delete_qr_code', methods=['POST'])
@admin_role_required
def delete_qr_code():
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    try:
        cursor.execute("SELECT value FROM settings WHERE key = 'store_qr_code_path'")
        result = cursor.fetchone()
        if result and result['value']:
            filepath = os.path.join(app.config['IMAGES_FOLDER'], result['value'])
            if os.path.exists(filepath): os.remove(filepath)
        cursor.execute("UPDATE settings SET value = '' WHERE key = 'store_qr_code_path'")
        db.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/update_opening', methods=['POST'])
@admin_role_required
def update_opening():
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    try:
        key_map = {'opening_message': 'opening_message', 'writing_mode': 'opening_writing_mode', 'opening_effect': 'opening_effect', 'opening_duration': 'opening_duration'}
        for form_key, db_key in key_map.items():
            if request.form.get(form_key) is not None:
                cursor.execute("UPDATE settings SET value = %s WHERE key = %s", (request.form.get(form_key), db_key))
        for db_key, file in [('opening_image_path', request.files.get('opening_image_1')), ('opening_image_path_2', request.files.get('opening_image_2'))]:
            if file and file.filename != '':
                img = Image.open(file.stream)
                ext = os.path.splitext(secure_filename(file.filename))[1].lower()
                if ext not in ['.png', '.jpg', '.jpeg', '.gif', '.webp']: raise IOError("Invalid image format")
                filename = f"{db_key.replace('_path', '')}{ext}"
                img.save(os.path.join(app.config['IMAGES_FOLDER'], filename))
                cursor.execute("UPDATE settings SET value = %s WHERE key = %s", (filename, db_key))
        db.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/delete_opening_image/<int:image_number>', methods=['POST'])
@admin_role_required
def delete_opening_image(image_number):
    if image_number not in [1, 2]: return jsonify({"status": "error", "message": "無効な画像番号です"}), 400
    db_key = 'opening_image_path' if image_number == 1 else 'opening_image_path_2'
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    try:
        cursor.execute("SELECT value FROM settings WHERE key = %s", (db_key,))
        result = cursor.fetchone()
        if result and result['value']:
            filepath = os.path.join(app.config['IMAGES_FOLDER'], result['value'])
            if os.path.exists(filepath): os.remove(filepath)
        cursor.execute("UPDATE settings SET value = '' WHERE key = %s", (db_key,))
        db.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/change_password', methods=['POST'])
@admin_role_required
def change_password():
    data = request.get_json()
    target_user, current_pass, new_pass = data.get('username'), data.get('current_password'), data.get('new_password')
    if not all([target_user, current_pass, new_pass, data.get('confirm_password')]): return jsonify({"status": "error", "message": "すべての項目を入力してください。"}), 400
    if new_pass != data.get('confirm_password'): return jsonify({"status": "error", "message": "新しいパスワードが一致しません。"}), 400
    if len(new_pass) < 6: return jsonify({"status": "error", "message": "新しいパスワードは6文字以上にしてください。"}), 400
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT * FROM users WHERE username = %s", (g.user_username,))
    operator_user = cursor.fetchone()
    if not operator_user or not bcrypt.check_password_hash(operator_user['password_hash'], current_pass):
        return jsonify({"status": "error", "message": "現在のあなたのパスワードが正しくありません。"}), 403
    cursor.execute("SELECT * FROM users WHERE username = %s", (target_user,))
    if not cursor.fetchone(): return jsonify({"status": "error", "message": "対象のユーザーが見つかりません。"}), 404
    try:
        new_hash = bcrypt.generate_password_hash(new_pass).decode('utf-8')
        cursor.execute("UPDATE users SET password_hash = %s WHERE username = %s", (new_hash, target_user))
        db.commit()
        return jsonify({"status": "success", "message": f"{target_user}のパスワードを変更しました。"})
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/get_sales_data', methods=['GET'])
@admin_role_required
def get_sales_data():
    start_date_str, end_date_str = request.args.get('start_date'), request.args.get('end_date')
    if not start_date_str or not end_date_str: return jsonify({"status": "error", "message": "日付が指定されていません"}), 400
    try:
        start_ts = datetime.strptime(start_date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc).timestamp()
        end_ts = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=timezone.utc).timestamp()
        cursor = get_db().cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT o.created_at, o.table_id, oi.item_name, oi.quantity, oi.price FROM orders o JOIN order_items oi ON o.id = oi.order_id WHERE o.status = 'paid' AND o.paid_at BETWEEN %s AND %s ORDER BY o.paid_at ASC", (start_ts, end_ts))
        return jsonify(cursor.fetchall())
    except Exception as e: return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/get_cooking_times', methods=['GET'])
@admin_role_required
def get_cooking_times():
    start_date_str, end_date_str = request.args.get('start_date'), request.args.get('end_date')
    if not start_date_str or not end_date_str: return jsonify({"status": "error", "message": "日付が指定されていません"}), 400
    try:
        start_ts = datetime.strptime(start_date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc).timestamp()
        end_ts = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=timezone.utc).timestamp()
        cursor = get_db().cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT oi.item_name, oi.ready_at - o.created_at AS cooking_duration FROM order_items oi JOIN orders o ON oi.order_id = o.id WHERE o.status = 'paid' AND o.paid_at BETWEEN %s AND %s AND oi.ready_at IS NOT NULL AND o.created_at IS NOT NULL", (start_ts, end_ts))
        product_times = {}
        for row in cursor.fetchall():
            if 0 <= row['cooking_duration'] < 86400:
                if row['item_name'] not in product_times: product_times[row['item_name']] = []
                product_times[row['item_name']].append(row['cooking_duration'])
        avg_cooking_times = {name: round((sum(d) / len(d)) / 60, 1) for name, d in product_times.items() if d}
        return jsonify(avg_cooking_times)
    except Exception as e: return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/get_session_durations', methods=['GET'])
@admin_role_required
def get_session_durations():
    start_date_str, end_date_str = request.args.get('start_date'), request.args.get('end_date')
    if not start_date_str or not end_date_str: return jsonify({"status": "error", "message": "日付が指定されていません"}), 400
    try:
        start_ts = datetime.strptime(start_date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc).timestamp()
        end_ts = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=timezone.utc).timestamp()
        cursor = get_db().cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT id, table_id, created_at, paid_at, total_price FROM orders WHERE status = 'paid' AND paid_at BETWEEN %s AND %s", (start_ts, end_ts))
        paid_orders = cursor.fetchall()
        session_durations = []
        for order in paid_orders:
            cursor.execute("SELECT created_at FROM table_sessions WHERE table_id = %s AND created_at < %s ORDER BY created_at DESC LIMIT 1", (order['table_id'], order['paid_at']))
            session = cursor.fetchone()
            start_time = session['created_at'] if session else order['created_at']
            session_durations.append({"table_id": order['table_id'], "start_time": start_time, "end_time": order['paid_at'], "duration_minutes": round((order['paid_at'] - start_time) / 60), "total_price": order['total_price']})
        return jsonify(session_durations)
    except Exception as e: return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/download_menu', methods=['GET'])
@admin_role_required
def download_menu():
    try:
        db = get_db()
        cursor = db.cursor(cursor_factory=RealDictCursor)
        workbook = openpyxl.Workbook()
        sheet_cat = workbook.active
        sheet_cat.title = "カテゴリー設定"
        sheet_cat.append(['表示順', 'カテゴリー名(日本語)', 'カテゴリー名(English)'])
        cursor.execute("SELECT display_order, name_jp, name_en FROM categories ORDER BY display_order, id")
        for cat in cursor.fetchall():
            sheet_cat.append([cat['display_order'], cat['name_jp'], cat['name_en']])
        sheet_menu = workbook.create_sheet(title="メニュー")
        cursor.execute("SELECT id, name_jp FROM categories")
        cat_map = {row['id']: row['name_jp'] for row in cursor.fetchall()}
        cursor.execute("SELECT product_id, category_id FROM product_categories")
        product_cat_map = {}
        for row in cursor.fetchall():
            pid, cid = row['product_id'], row['category_id']
            if pid not in product_cat_map: product_cat_map[pid] = []
            if cid in cat_map: product_cat_map[pid].append(cat_map[cid])
        cursor.execute("SELECT * FROM products ORDER BY id")
        products = cursor.fetchall()
        headers = ['商品ID', '商品名', '価格', '商品説明', '画像ファイル名', 'カテゴリー', '品切れ', '商品名(英語)', '商品説明(英語)']
        sheet_menu.append(headers)
        for product in products:
            category_names = " ".join(product_cat_map.get(product['id'], []))
            sheet_menu.append([product['id'], product['name'], product['price'], product['description'], product['image_path'], category_names, product['is_sold_out'], product['name_en'], product['description_en']])
        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"menu_backup_{timestamp}.xlsx"
        return send_file(file_stream, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"status": "error", "message": f"ダウンロードファイルの生成に失敗しました: {str(e)}"}), 500

@app.route('/api/admin/upload_menu', methods=['POST'])
@admin_role_required
def upload_menu():
    if 'menu_file' not in request.files: return jsonify({"status": "error", "message": "ファイルがありません"}), 400
    file = request.files['menu_file']
    if file.filename == '' or not file.filename.endswith('.xlsx'): return jsonify({"status": "error", "message": "無効なファイル形式です。"}), 400
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(filepath)
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    try:
        workbook = openpyxl.load_workbook(filepath)
        if "カテゴリー設定" in workbook.sheetnames:
            sheet_cat = workbook["カテゴリー設定"]
            cursor.execute("DELETE FROM product_categories")
            cursor.execute("DELETE FROM categories")
            cursor.execute("ALTER SEQUENCE categories_id_seq RESTART WITH 1")
            categories_to_insert = []
            for row in sheet_cat.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]: continue
                order, name_jp, name_en = (row + (99, None, None))[:3]
                categories_to_insert.append((int(order or 99), name_jp, name_en))
            for cat in categories_to_insert:
                cursor.execute("INSERT INTO categories (display_order, name_jp, name_en) VALUES (%s, %s, %s)", cat)

        count = 0
        if "メニュー" in workbook.sheetnames:
            sheet_menu = workbook["メニュー"]
            cursor.execute("SELECT id, name_jp FROM categories")
            cat_name_to_id_map = {row['name_jp']: row['id'] for row in cursor.fetchall()}
            cursor.execute("DELETE FROM products")
            cursor.execute("ALTER SEQUENCE products_id_seq RESTART WITH 1")
            
            for row_index, row in enumerate(sheet_menu.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[1]: continue
                _id, name, price, desc, img, cat_str, sold_out, name_en, desc_en = (row + (None,) * 9)[:9]

            # ★★★ ここから追記 ★★★
            # もし画像ファイル名(img)が存在すれば、小文字に変換する
            if img:
                img = str(img).lower()
            # ★★★ ここまで追記 ★★★


                price_val = 0
                if price is not None:
                    if isinstance(price, str):
                        cleaned_price = re.sub(r'[^\d]', '', price)
                        if cleaned_price: price_val = int(cleaned_price)
                    elif isinstance(price, (int, float)):
                        price_val = int(price)
                
                cursor.execute("INSERT INTO products (name, price, description, image_path, is_sold_out, name_en, description_en) VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id", (name, price_val, desc, img, 1 if sold_out else 0, name_en, desc_en))
                new_product_id = cursor.fetchone()['id']
                
                if cat_str:
                    category_names = cat_str.replace('　', ' ').split(' ')
                    for cat_name in [c for c in category_names if c]:
                        cat_id = cat_name_to_id_map.get(cat_name.strip())
                        if cat_id:
                            cursor.execute("INSERT INTO product_categories (product_id, category_id) VALUES (%s, %s) ON CONFLICT DO NOTHING", (new_product_id, cat_id))
                count += 1
        db.commit()
        return jsonify({"status": "success", "message": f"{count}件のメニューとカテゴリーを登録/更新しました。"})
    except Exception as e:
        db.rollback()
        error_info = f"エラーが発生しました: {str(e)}"
        if 'row_index' in locals():
            error_info += f" (Excelの {row_index} 行目付近)"
        return jsonify({"status": "error", "message": error_info}), 500
    finally:
        if os.path.exists(filepath): os.remove(filepath)

@app.route('/api/admin/add_product', methods=['POST'])
@admin_role_required
def add_product():
    # ...
    try:
        # ...
        image_file = request.files.get('image_file')
        # ...
        if image_file:
            img = Image.open(image_file.stream)
            # ★★★ ここから修正 ★★★
            original_filename = secure_filename(image_file.filename)
            filename_without_ext, ext = os.path.splitext(original_filename)
            # 新しいファイル名を生成し、拡張子も含めて全て小文字に変換する
            new_filename = f"{product_id}{ext}".lower() 
            # ★★★ ここまで修正 ★★★
            img.save(os.path.join(app.config['IMAGES_FOLDER'], new_filename))
            cursor.execute("UPDATE products SET image_path = %s WHERE id = %s", (new_filename, product_id))


        if category_ids:
            for cid in category_ids:
                cursor.execute("INSERT INTO product_categories (product_id, category_id) VALUES (%s, %s)", (product_id, cid))

        db.commit()
        return jsonify({"status": "success", "productId": product_id})
    except psycopg2.errors.UniqueViolation:
        db.rollback()
        return jsonify({"status": "error", "message": "その品名は既に使用されています。"}), 409
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/update_product/<int:product_id>', methods=['POST'])
@admin_role_required
def update_product(product_id):
    data = request.get_json()
    # ...
    try:
        # ...
        # ★★★ ここから修正 ★★★
        # image_pathが存在すれば、小文字に変換してからDBに保存する
        image_path = data.get('image_path')
        if image_path:
            image_path = image_path.lower()
        # ★★★ ここまで修正 ★★★
        cursor.execute("UPDATE products SET name = %s, price = %s, description = %s, name_en = %s, description_en = %s, image_path = %s WHERE id = %s", (data['name'], data['price'], data['description'], data['name_en'], data['description_en'], image_path, product_id))
        cursor.execute("DELETE FROM product_categories WHERE product_id = %s", (product_id,))
        if category_ids:
            for cid in category_ids:
                cursor.execute("INSERT INTO product_categories (product_id, category_id) VALUES (%s, %s)", (product_id, cid))
        db.commit()
        return jsonify({"status": "success"})
    except psycopg2.errors.UniqueViolation:
        db.rollback()
        return jsonify({"status": "error", "message": "その品名は既に使用されています。"}), 409
    except Exception as e:
        db.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/update_product_status/<int:product_id>', methods=['POST'])
@admin_role_required
def update_product_status(product_id):
    data = request.get_json()
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("UPDATE products SET is_sold_out = %s WHERE id = %s", (data['is_sold_out'], product_id))
    db.commit()
    return jsonify({"status": "success"})
        
@app.route('/api/admin/delete_product/<int:product_id>', methods=['POST'])
@admin_role_required
def delete_product(product_id):
    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)
    cursor.execute("DELETE FROM products WHERE id = %s", (product_id,))
    db.commit()
    return jsonify({"status": "success"})


# --- 静的ファイル配信 ---
@app.route('/')
def serve_root():
    return send_from_directory('.', 'login.html')

@app.route('/favicon.ico')
def favicon():
    return '', 204

# --- ここから追記 ---
@app.route('/ROS_manual/<path:path>')
def serve_manual_files(path):
    return send_from_directory('ROS_manual', path)
# --- ここまで追記 ---

@app.route('/<path:path>')
def serve_static_file(path):
    if '..' in path: return "Not Found", 404
    safe_path = os.path.join('.', path)
    if os.path.exists(safe_path) and not os.path.isdir(safe_path):
        return send_from_directory('.', path)
    return "Not Found", 404

# --- データベース初期化コマンド ---
@app.cli.command("init-db")
def init_db_command():
    """データベーステーブルを作成し、マイグレーションを実行します。"""
    init_db()
    migrate_db()
    print("Initialized and migrated the database.")

# --- 実行 ---
# --- 実行 ---
if __name__ == '__main__':
    # Renderが指定する環境変数'PORT'を取得。なければ開発用に5000を使う
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)